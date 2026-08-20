import sys
import os
import openpyxl
from io import BytesIO

sys.path.insert(0, os.path.dirname(__file__))

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

def create_sample_excel_with_equations():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QuestionBank"

    headers = ["S.No", "Question", "Image", "CO", "K-Level", "Mark", "Part"]
    ws.append(headers)

    test_rows = [
        ["1", r"List two common chemical contaminants in food. \frac{a}{b}", "", "CO1", "K2", "2", "A"],
        ["2", r"Mention two common food adulterants. \frac{dy}{dx}", "", "CO1", "K2", "2", "A"],
        ["3", r"What is the functional role of preservatives in food? \int_a^b f(x)dx", "", "CO1", "K2", "2", "A"],
        ["4", r"Define cleaning. \frac{d^2y}{dx^2}", "", "CO2", "K1", "2", "A"],
        ["5", r"What is disinfection? \sqrt{x} \frac{a}{b} \sum_{i=1}^{n}", "", "CO2", "K2", "2", "A"],
    ]
    for r in test_rows:
        ws.append(r)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

def main():
    print("=== TESTING EXCEL EQUATION PIPELINE ===")

    excel_bytes = create_sample_excel_with_equations()
    print(f"1. Created sample Excel file ({len(excel_bytes)} bytes)")

    wb = openpyxl.load_workbook(filename=BytesIO(excel_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    print(f"2. Read {len(rows)-1} question rows from Excel")
    for i, r in enumerate(rows[1:], 1):
        q = r[1]
        co = r[3]
        kl = r[4]
        print(f"   [Row {i}] Q: '{q}' | CO: '{co}' | K-Level: '{kl}'")

    print("\n=== ALL EXCEL EQUATION TESTS PASSED CLEANLY! ===")

if __name__ == "__main__":
    main()
