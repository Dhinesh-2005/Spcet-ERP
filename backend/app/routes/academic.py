import re
import openpyxl
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any, Optional
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from bson import ObjectId
from app.database import db, clean_doc, clean_docs
from app.models.schemas import (
    DepartmentModel, RegulationModel, AcademicYearModel, SubjectModel,
    FacultySubjectAccessModel, AssessmentConfigModel, PreviousResultRow
)
from app.services.academic_service import academic_service, GRADE_MAP

router = APIRouter()

def _clean_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()

# ─────────────────────────────────────────────────────────────────────────────
# 1. DEPARTMENTS, REGULATIONS, ACADEMIC YEARS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/departments")
async def get_departments():
    docs = [clean_doc(d) async for d in db.departments.find({})]
    if not docs:
        # Default seed if empty
        default_depts = [
            {"code": "CSE", "name": "Computer Science and Engineering"},
            {"code": "IT", "name": "Information Technology"},
            {"code": "ECE", "name": "Electronics and Communication Engineering"},
            {"code": "EEE", "name": "Electrical and Electronics Engineering"},
            {"code": "AIDS", "name": "Artificial Intelligence and Data Science"},
            {"code": "MECH", "name": "Mechanical Engineering"},
            {"code": "CIVIL", "name": "Civil Engineering"},
            {"code": "BME", "name": "Biomedical Engineering"},
            {"code": "CSBS", "name": "Computer Science and Business Systems"},
            {"code": "BIOTECH", "name": "Biotechnology"},
            {"code": "AERO", "name": "Aeronautical Engineering"}
        ]
        return default_depts
    return docs

@router.post("/departments")
async def create_department(dept: DepartmentModel):
    dept_code = dept.code.strip().upper()
    existing = await db.departments.find_one({"code": dept_code})
    if existing:
        raise HTTPException(status_code=400, detail=f"Department '{dept_code}' already exists.")
    doc = dept.dict(exclude_none=True, exclude={"id"})
    doc["code"] = dept_code
    res = await db.departments.insert_one(doc)
    doc["id"] = str(res.inserted_id)
    return {"message": "Department created", "department": doc}

@router.get("/regulations")
async def get_regulations():
    docs = [clean_doc(d) async for d in db.regulations.find({})]
    if not docs:
        return [
            {"code": "2021", "name": "Regulation 2021", "effectiveYear": 2021},
            {"code": "2024", "name": "Regulation 2024", "effectiveYear": 2024}
        ]
    return docs

@router.post("/regulations")
async def create_regulation(reg: RegulationModel):
    reg_code = reg.code.strip()
    existing = await db.regulations.find_one({"code": reg_code})
    if existing:
        raise HTTPException(status_code=400, detail=f"Regulation '{reg_code}' already exists.")
    doc = reg.dict(exclude_none=True, exclude={"id"})
    res = await db.regulations.insert_one(doc)
    doc["id"] = str(res.inserted_id)
    return {"message": "Regulation created", "regulation": doc}

@router.get("/academic-years")
async def get_academic_years():
    docs = [clean_doc(d) async for d in db.academic_years.find({})]
    if not docs:
        return [
            {"year": "2023-2024", "isCurrent": False},
            {"year": "2024-2025", "isCurrent": False},
            {"year": "2025-2026", "isCurrent": False},
            {"year": "2026-2027", "isCurrent": True}
        ]
    return docs

# ─────────────────────────────────────────────────────────────────────────────
# 2. SUBJECT MASTER
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/subjects")
async def get_subjects(
    department: Optional[str] = Query(None),
    semester: Optional[int] = Query(None),
    paperType: Optional[str] = Query(None),
    regulation: Optional[str] = Query(None)
):
    query = {}
    if department and department != "ALL":
        query["department"] = department.strip().upper()
    if semester:
        query["semester"] = int(semester)
    if paperType and paperType != "ALL":
        query["paperType"] = paperType.strip().upper()
    if regulation and regulation != "ALL":
        query["regulation"] = regulation.strip()

    cursor = db.subjects.find(query).sort("subjectCode", 1)
    return [clean_doc(d) async for d in cursor]

@router.post("/subjects")
async def add_or_update_subject(sub: SubjectModel):
    sub.subjectCode = sub.subjectCode.strip().upper()
    sub.department = sub.department.strip().upper()
    if not sub.subjectCode or not sub.department:
        raise HTTPException(status_code=400, detail="Subject code and department are required.")

    doc = sub.dict(exclude_none=True, exclude={"id"})
    await db.subjects.update_one(
        {"subjectCode": sub.subjectCode, "department": sub.department},
        {"$set": doc},
        upsert=True
    )
    return {"message": f"Subject {sub.subjectCode} saved successfully."}

@router.post("/subjects/upload-excel")
async def upload_subjects_excel(
    file: UploadFile = File(...),
    regulation: str = Form("2021"),
    defaultSemester: Optional[int] = Form(None)
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported.")

    content = await file.read()
    try:
        workbook = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")

    total_imported = 0
    total_updated = 0
    dept_summary = {}
    skipped_details = []

    for sheet in workbook.worksheets:
        dept_name = sheet.title.strip().upper()
        # Skip hidden/meta sheets if title starts with _
        if dept_name.startswith("_"):
            continue

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        header_idx = -1
        for idx, row in enumerate(rows[:15]):
            if not row:
                continue
            row_str = " ".join([_clean_str(c).lower() for c in row if c])
            if "subject" in row_str or "code" in row_str or "paper" in row_str or "credit" in row_str:
                header_idx = idx
                break

        if header_idx == -1:
            header_idx = 0

        header_cells = [_clean_str(c) for c in (rows[header_idx] or [])]
        header_lower = [c.lower() for c in header_cells]

        def _col_idx(candidates):
            for idx, cell in enumerate(header_lower):
                clean_cell = re.sub(r"[^a-z0-9]", "", cell)
                for cand in candidates:
                    if cand in clean_cell:
                        return idx
            return -1

        sub_code_idx = _col_idx(["subjectcode", "subcode", "code"])
        sub_name_idx = _col_idx(["subjectname", "subname", "name", "title", "subjecttitle"])
        credits_idx = _col_idx(["credits", "credit"])
        type_idx = _col_idx(["papertype", "type", "paper", "category"])
        sem_idx = _col_idx(["semester", "sem"])

        if sub_code_idx == -1 and len(header_cells) >= 1:
            sub_code_idx = 0

        if sub_code_idx == -1:
            skipped_details.append(f"Sheet '{sheet.title}': No Subject Code column found.")
            continue

        sheet_processed = 0
        for row in rows[header_idx + 1:]:
            if not row or not any(row):
                continue

            def cell_val(idx):
                if idx != -1 and idx < len(row):
                    return _clean_str(row[idx])
                return ""

            code = cell_val(sub_code_idx).upper()
            if not code or len(code) < 3:
                continue

            name = cell_val(sub_name_idx) or code
            
            raw_credits = cell_val(credits_idx)
            try:
                credits = float(raw_credits) if raw_credits else 0.0
            except ValueError:
                credits = 0.0

            paper_type = cell_val(type_idx).upper() or "THEORY"
            if "PRAC" in paper_type or "LAB" in paper_type:
                paper_type = "PRACTICAL"
            elif "INTEG" in paper_type:
                paper_type = "INTEGRATED"
            elif "ELEC" in paper_type:
                paper_type = "ELECTIVE"
            elif "PROJ" in paper_type:
                paper_type = "PROJECT"
            elif not paper_type:
                paper_type = "THEORY"

            raw_sem = cell_val(sem_idx)
            try:
                sem = int(float(raw_sem)) if raw_sem else int(defaultSemester)
            except ValueError:
                sem = int(defaultSemester)

            doc = {
                "subjectCode": code,
                "subjectName": name,
                "department": dept_name,
                "semester": sem,
                "credits": credits,
                "paperType": paper_type,
                "regulation": regulation.strip()
            }

            res = await db.subjects.update_one(
                {"subjectCode": code, "department": dept_name, "regulation": regulation.strip()},
                {"$set": doc},
                upsert=True
            )

            if res.upserted_id:
                total_imported += 1
            else:
                total_updated += 1

            sheet_processed += 1

        if sheet_processed > 0:
            dept_summary[dept_name] = sheet_processed

    return {
        "message": f"Successfully processed {sum(dept_summary.values())} subjects across {len(dept_summary)} department sheet(s).",
        "totalImported": total_imported,
        "totalUpdated": total_updated,
        "departmentSummary": dept_summary,
        "skippedDetails": skipped_details
    }


@router.get("/subjects/sample-template")
async def download_subject_sample_template():
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default Sheet

    sample_departments = {
        "CSE": [
            ("CS3501", "Compiler Design", 4, "THEORY"),
            ("CS3591", "Computer Networks Lab", 2, "PRACTICAL"),
            ("CS3551", "Artificial Intelligence", 3, "THEORY"),
            ("CS3601", "Machine Learning", 4, "INTEGRATED"),
            ("CS3691", "Machine Learning Lab", 2, "PRACTICAL"),
            ("CS3602", "Cloud Computing", 3, "THEORY"),
        ],
        "ECE": [
            ("EC3501", "Wireless Communication", 4, "THEORY"),
            ("EC3591", "Communication Systems Lab", 2, "PRACTICAL"),
            ("EC3551", "VLSI Design", 3, "THEORY"),
            ("EC3601", "Embedded Systems", 4, "INTEGRATED"),
        ],
        "IT": [
            ("IT3501", "Web Technology", 3, "THEORY"),
            ("IT3591", "Web Technology Lab", 2, "PRACTICAL"),
            ("IT3601", "Cyber Security", 3, "THEORY"),
        ],
        "AIDS": [
            ("AD3501", "Deep Learning", 4, "THEORY"),
            ("AD3591", "Data Analytics Lab", 2, "PRACTICAL"),
            ("AD3601", "Natural Language Processing", 4, "INTEGRATED"),
        ]
    }

    for dept, subjects in sample_departments.items():
        ws = wb.create_sheet(title=dept)
        ws.append(["Subject Code", "Subject Name", "Credits", "Paper Type"])
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        for row in subjects:
            ws.append(list(row))

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=subject_upload_template.xlsx"}
    )

@router.post("/subjects/bulk-delete")
async def bulk_delete_subjects(payload: Dict[str, Any]):
    subject_ids = payload.get("subjectIds", [])
    subject_keys = payload.get("subjectKeys", [])

    deleted_count = 0

    if subject_ids:
        obj_ids = []
        for sid in subject_ids:
            try:
                obj_ids.append(ObjectId(sid))
            except Exception:
                pass
        if obj_ids:
            subs = [clean_doc(d) async for d in db.subjects.find({"_id": {"$in": obj_ids}})]
            codes = list(set([s.get("subjectCode") for s in subs if s.get("subjectCode")]))
            
            res = await db.subjects.delete_many({"_id": {"$in": obj_ids}})
            deleted_count += res.deleted_count
            if codes:
                await db.student_subjects.delete_many({"subjectCode": {"$in": codes}})

    if subject_keys:
        for key in subject_keys:
            code = _clean_str(key.get("subjectCode")).upper()
            dept = _clean_str(key.get("department")).upper()
            reg = _clean_str(key.get("regulation"))

            q = {}
            if code: q["subjectCode"] = code
            if dept: q["department"] = dept
            if reg: q["regulation"] = reg

            if q:
                res = await db.subjects.delete_many(q)
                deleted_count += res.deleted_count
                if code:
                    await db.student_subjects.delete_many({"subjectCode": code})

    return {
        "message": f"Successfully deleted {deleted_count} subject(s).",
        "deletedCount": deleted_count
    }

@router.delete("/subjects/{subjectCode}")
async def delete_single_subject(subjectCode: str, department: Optional[str] = Query(None), regulation: Optional[str] = Query(None)):
    code = subjectCode.strip().upper()
    q = {"subjectCode": code}
    if department:
        q["department"] = department.strip().upper()
    if regulation:
        q["regulation"] = regulation.strip()

    result = await db.subjects.delete_one(q)
    await db.student_subjects.delete_many({"subjectCode": code})

    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail=f"Subject {subjectCode} not found.")

    return {"message": f"Subject {subjectCode} deleted successfully."}

# ─────────────────────────────────────────────────────────────────────────────
# 3. FACULTY SUBJECT ACCESS (Permissions)
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/faculty-access")
async def get_all_faculty_access():
    cursor = db.faculty_subject_access.find({})
    return [clean_doc(d) async for d in cursor]

@router.get("/faculty-access/{faculty_id}")
async def get_faculty_access(faculty_id: str):
    fac_id = faculty_id.strip()
    doc = await db.faculty_subject_access.find_one({"facultyId": fac_id})
    if not doc:
        # Fallback check in faculties collection
        fac = await db.faculties.find_one({"registerNumber": fac_id})
        return {
            "facultyId": fac_id,
            "facultyName": fac.get("name", "") if fac else "",
            "department": fac.get("department", "") if fac else "",
            "subjectCodes": []
        }
    return clean_doc(doc)

@router.post("/faculty-access")
async def save_faculty_access(payload: FacultySubjectAccessModel):
    fac_id = payload.facultyId.strip()
    if not fac_id:
        raise HTTPException(status_code=400, detail="Faculty ID is required.")

    doc = payload.dict(exclude_none=True, exclude={"id"})
    doc["facultyId"] = fac_id

    await db.faculty_subject_access.update_one(
        {"facultyId": fac_id},
        {"$set": doc},
        upsert=True
    )
    return {"message": f"Subject access updated for faculty {fac_id}."}

# ─────────────────────────────────────────────────────────────────────────────
# 4. ASSESSMENT CONFIG (Unit Tests / Assignments Max Marks Scheme)
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/assessment-config")
async def get_assessment_config(
    department: str = Query("CSE"),
    semester: int = Query(1),
    subjectCode: Optional[str] = Query(None)
):
    query = {"department": department.strip().upper(), "semester": int(semester)}
    if subjectCode:
        query["subjectCode"] = subjectCode.strip().upper()

    doc = await db.assessment_configs.find_one(query)
    if not doc:
        # Default component scheme if not configured yet
        default_components = [
            {"name": "UT1", "maxMarks": 20.0, "weightage": 1.0},
            {"name": "UT2", "maxMarks": 20.0, "weightage": 1.0},
            {"name": "Unit 1", "maxMarks": 20.0, "weightage": 1.0},
            {"name": "Unit 2", "maxMarks": 20.0, "weightage": 1.0},
            {"name": "Unit 3", "maxMarks": 20.0, "weightage": 1.0},
            {"name": "Unit 4", "maxMarks": 20.0, "weightage": 1.0},
            {"name": "Unit 5", "maxMarks": 20.0, "weightage": 1.0},
            {"name": "Assignment", "maxMarks": 10.0, "weightage": 1.0},
        ]
        return {
            "department": department,
            "semester": semester,
            "subjectCode": subjectCode,
            "components": default_components,
            "totalInternalMaxMark": 40.0
        }
    return clean_doc(doc)

@router.post("/assessment-config")
async def save_assessment_config(payload: AssessmentConfigModel):
    query = {
        "department": payload.department.strip().upper(),
        "semester": int(payload.semester)
    }
    if payload.subjectCode:
        query["subjectCode"] = payload.subjectCode.strip().upper()
    else:
        query["subjectCode"] = None

    doc = payload.dict(exclude_none=True, exclude={"id"})
    doc["department"] = payload.department.strip().upper()

    await db.assessment_configs.update_one(
        query,
        {"$set": doc},
        upsert=True
    )
    return {"message": "Assessment configuration saved successfully."}

# ─────────────────────────────────────────────────────────────────────────────
# 5. PREVIOUS SEMESTER RESULTS EXCEL IMPORT (With Validation & Preview)
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/previous-results/validate")
async def validate_previous_results_excel(
    file: UploadFile = File(...),
    regulation: Optional[str] = Form(None),
    department: Optional[str] = Form(None),
    studentCurrentSem: Optional[int] = Form(None),
    resultSem: Optional[int] = Form(None)
):
    content = await file.read()
    try:
        workbook = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    header_idx = -1
    for idx, row in enumerate(rows[:15]):
        if not row:
            continue
        row_str = " ".join([_clean_str(c).lower() for c in row if c])
        if "reg" in row_str or "register" in row_str or "subject" in row_str or "grade" in row_str or "roll" in row_str:
            header_idx = idx
            break
    if header_idx == -1:
        header_idx = 0

    header_cells = [_clean_str(c) for c in (rows[header_idx] or [])]
    header_lower = [c.lower() for c in header_cells]

    # Pre-fetch all subject master records into a dictionary: sub_code -> doc
    db_subjects = {}
    async for sdoc in db.subjects.find({}):
        code = _clean_str(sdoc.get("subjectCode")).upper()
        if code:
            db_subjects[code] = sdoc

    def _col(names):
        for idx, cell in enumerate(header_lower):
            clean_cell = re.sub(r"[^a-z0-9]", "", cell)
            for n in names:
                if n in clean_cell:
                    return idx
        return -1

    reg_idx = _col(["registernumber", "regno", "register", "rollno", "rollnumber"])
    sub_code_idx = _col(["subjectcode", "subcode", "code"])
    grade_idx = _col(["grade"])

    is_matrix_format = False
    if reg_idx != -1 and (sub_code_idx == -1 or grade_idx == -1):
        # Check if remaining headers look like subject codes
        potential_subject_cols = []
        for c_idx in range(len(header_cells)):
            if c_idx == reg_idx:
                continue
            h_text = header_cells[c_idx].strip()
            if h_text and len(h_text) >= 3:
                potential_subject_cols.append((c_idx, h_text.upper()))
        if potential_subject_cols:
            is_matrix_format = True

    parsed_rows = []
    errors = []
    duplicates = []
    seen_keys = set()

    sem_val = resultSem or 1

    if is_matrix_format:
        reg_col = reg_idx if reg_idx != -1 else 0
        subject_columns = []
        for c_idx in range(len(header_cells)):
            if c_idx == reg_col:
                continue
            h_text = header_cells[c_idx].strip()
            if h_text:
                subject_columns.append((c_idx, h_text.upper()))

        for row_num in range(header_idx + 1, len(rows)):
            row = rows[row_num]
            if not row or len(row) <= reg_col:
                continue

            reg_no = _clean_str(row[reg_col]).upper()
            if not reg_no or len(reg_no) < 3:
                continue

            for c_idx, sub_code in subject_columns:
                if c_idx >= len(row):
                    continue
                cell_val = row[c_idx]
                if cell_val is None or _clean_str(cell_val) == "":
                    continue

                mark_or_grade = _clean_str(cell_val).upper()

                # Determine Grade and Grade Point from numerical mark or grade string
                grade = "U"
                gp_val = 0.0
                res_val = "FAIL"

                # If numerical mark
                try:
                    num_mark = float(mark_or_grade)
                    if num_mark >= 90:
                        grade, gp_val, res_val = "O", 10.0, "PASS"
                    elif num_mark >= 80:
                        grade, gp_val, res_val = "A+", 9.0, "PASS"
                    elif num_mark >= 70:
                        grade, gp_val, res_val = "A", 8.0, "PASS"
                    elif num_mark >= 60:
                        grade, gp_val, res_val = "B+", 7.0, "PASS"
                    elif num_mark >= 50:
                        grade, gp_val, res_val = "B", 6.0, "PASS"
                    else:
                        grade, gp_val, res_val = "U", 0.0, "FAIL"
                except ValueError:
                    grade_info = academic_service.get_grade_info(mark_or_grade)
                    grade = mark_or_grade
                    gp_val = grade_info["gp"]
                    res_val = grade_info["result"]

                # Fetch credits & name from Subject Master
                sub_doc = db_subjects.get(sub_code, {})
                credits_val = int(sub_doc.get("credits") or 3)
                sub_name = sub_doc.get("subjectName") or sub_code

                key = (reg_no, sub_code, sem_val)
                if key in seen_keys:
                    duplicates.append(f"Duplicate entry in file for student {reg_no}, subject {sub_code}, semester {sem_val}")
                seen_keys.add(key)

                existing = await db.student_subjects.find_one({
                    "registerNumber": reg_no,
                    "subjectCode": sub_code,
                    "semester": sem_val
                })
                is_existing_db = True if existing else False
                if is_existing_db:
                    duplicates.append(f"Database already contains record for {reg_no} - {sub_code} (Semester {sem_val})")

                parsed_rows.append({
                    "registerNumber": reg_no,
                    "subjectCode": sub_code,
                    "subjectName": sub_name,
                    "semester": sem_val,
                    "academicYear": "2025-2026",
                    "credits": credits_val,
                    "grade": grade,
                    "gradePoint": gp_val,
                    "result": res_val,
                    "department": department or sub_doc.get("department") or "CSE",
                    "studentCurrentSem": studentCurrentSem,
                    "isExistingInDb": is_existing_db
                })
    else:
        # Standard tabular format
        sub_name_idx = _col(["subjectname", "subname", "subjecttitle", "title"])
        sem_idx = _col(["semester", "sem"])
        ay_idx = _col(["academicyear", "academicyr", "batch", "year"])
        credit_idx = _col(["credits", "credit"])
        gp_idx = _col(["gradepoint", "gp"])
        result_idx = _col(["result", "status", "passfail"])

        if reg_idx == -1 or sub_code_idx == -1 or grade_idx == -1:
            raise HTTPException(
                status_code=400,
                detail="Excel missing required columns. Ensure 'Register Number', 'Subject Code', and 'Grade' are present (or matrix columns with subject headers)."
            )

        for row_num in range(header_idx + 1, len(rows)):
            row = rows[row_num]
            if not row:
                continue

            reg_no = _clean_str(row[reg_idx]).upper()
            if not reg_no:
                continue

            sub_code = _clean_str(row[sub_code_idx]).upper()
            grade = _clean_str(row[grade_idx]).upper()

            if not sub_code:
                errors.append(f"Row {row_num + 1}: Missing subject code for student {reg_no}")
                continue

            row_sem = sem_val
            if sem_idx != -1 and sem_idx < len(row) and row[sem_idx] is not None:
                try:
                    row_sem = int(row[sem_idx])
                except (ValueError, TypeError):
                    row_sem = sem_val

            sub_doc = db_subjects.get(sub_code, {})

            credits_val = int(sub_doc.get("credits") or 3)
            if credit_idx != -1 and credit_idx < len(row) and row[credit_idx] is not None:
                try:
                    credits_val = int(float(str(row[credit_idx])))
                except (ValueError, TypeError):
                    pass

            sub_name = _clean_str(row[sub_name_idx]) if sub_name_idx != -1 and sub_name_idx < len(row) else (sub_doc.get("subjectName") or sub_code)
            ay_val = _clean_str(row[ay_idx]) if ay_idx != -1 and ay_idx < len(row) else "2025-2026"

            grade_info = academic_service.get_grade_info(grade)
            gp_val = grade_info["gp"]
            res_val = grade_info["result"]

            if gp_idx != -1 and gp_idx < len(row) and row[gp_idx] is not None:
                try:
                    gp_val = float(row[gp_idx])
                except (ValueError, TypeError):
                    pass

            if result_idx != -1 and result_idx < len(row) and row[result_idx] is not None:
                r_str = _clean_str(row[result_idx]).upper()
                if r_str in ["PASS", "P"]:
                    res_val = "PASS"
                elif r_str in ["FAIL", "F", "RA"]:
                    res_val = "FAIL"

            key = (reg_no, sub_code, row_sem)
            if key in seen_keys:
                duplicates.append(f"Duplicate entry in file for student {reg_no}, subject {sub_code}, semester {row_sem}")
            seen_keys.add(key)

            existing = await db.student_subjects.find_one({
                "registerNumber": reg_no,
                "subjectCode": sub_code,
                "semester": row_sem
            })
            is_existing_db = True if existing else False
            if is_existing_db:
                duplicates.append(f"Database already contains record for {reg_no} - {sub_code} (Semester {row_sem})")

            parsed_rows.append({
                "registerNumber": reg_no,
                "subjectCode": sub_code,
                "subjectName": sub_name,
                "semester": row_sem,
                "academicYear": ay_val,
                "credits": credits_val,
                "grade": grade,
                "gradePoint": gp_val,
                "result": res_val,
                "department": department or sub_doc.get("department") or "CSE",
                "studentCurrentSem": studentCurrentSem,
                "isExistingInDb": is_existing_db
            })

    return {
        "totalRows": len(parsed_rows),
        "validRows": len(parsed_rows) - len(errors),
        "errorCount": len(errors),
        "errors": errors,
        "duplicateCount": len(duplicates),
        "duplicates": duplicates,
        "preview": parsed_rows
    }

@router.post("/previous-results/import")
async def import_previous_results(payload: Dict[str, Any]):
    rows = payload.get("rows", [])
    if not rows:
        raise HTTPException(status_code=400, detail="No rows provided for import.")

    imported = 0
    updated = 0

    for item in rows:
        reg_no = _clean_str(item.get("registerNumber")).upper()
        sub_code = _clean_str(item.get("subjectCode")).upper()
        if not reg_no or not sub_code:
            continue

        sem_val = int(item.get("semester") or 1)
        credits_val = int(item.get("credits") or 3)
        grade = _clean_str(item.get("grade")).upper()
        grade_info = academic_service.get_grade_info(grade)
        gp_val = float(item.get("gradePoint") if item.get("gradePoint") is not None else grade_info["gp"])
        res_val = item.get("result") or grade_info["result"]
        sub_name = item.get("subjectName") or sub_code
        ay_val = item.get("academicYear") or "2025-2026"
        dept_val = item.get("department") or "CSE"
        student_curr_sem = item.get("studentCurrentSem")

        doc = {
            "registerNumber": reg_no,
            "subjectCode": sub_code,
            "subjectName": sub_name,
            "semester": sem_val,
            "academicYear": ay_val,
            "credits": credits_val,
            "subjectType": "THEORY",
            "internalMark": 0.0,
            "externalMark": 0.0,
            "totalMark": gp_val * 10,
            "grade": grade,
            "gradePoint": gp_val,
            "result": res_val,
            "attemptNumber": 1,
            "category": "REGULAR",
            "updatedAt": datetime.utcnow().isoformat()
        }

        res = await db.student_subjects.update_one(
            {
                "registerNumber": reg_no,
                "subjectCode": sub_code,
                "semester": sem_val
            },
            {"$set": doc},
            upsert=True
        )

        if res.upserted_id:
            imported += 1
        else:
            updated += 1

        # Update or create student profile record to guarantee immediate visibility
        student_doc = await db.students.find_one({"registerNumber": reg_no})
        student_update = {}
        if dept_val and (not student_doc or student_doc.get("department") == "GENERAL"):
            student_update["department"] = dept_val
        if student_curr_sem:
            student_update["semester"] = int(student_curr_sem)
        elif not student_doc:
            student_update["semester"] = sem_val

        if not student_doc:
            await db.students.insert_one({
                "registerNumber": reg_no,
                "name": f"Student {reg_no}",
                "password": "password",
                "department": dept_val,
                "semester": int(student_curr_sem) if student_curr_sem else sem_val,
                "role": "student"
            })
        elif student_update:
            await db.students.update_one({"registerNumber": reg_no}, {"$set": student_update})

    return {
        "message": f"Previous semester results imported successfully.",
        "imported": imported,
        "updated": updated
    }

@router.get("/previous-results/count")
async def get_previous_results_count(
    department: Optional[str] = Query(None),
    resultSem: Optional[int] = Query(None),
    regulation: Optional[str] = Query(None),
    registerNumber: Optional[str] = Query(None)
):
    query = {}
    if resultSem is not None and resultSem != 0:
        query["semester"] = int(resultSem)
    if registerNumber:
        query["registerNumber"] = registerNumber.strip().upper()

    student_query = {}
    if department and department.upper() != "ALL":
        student_query["department"] = department.strip().upper()

    if student_query or (department and department.upper() != "ALL"):
        dept_students = [
            doc["registerNumber"]
            async for doc in db.students.find(student_query, {"registerNumber": 1})
        ]
        if registerNumber and registerNumber.strip().upper() not in dept_students:
            dept_students.append(registerNumber.strip().upper())

        if query.get("registerNumber"):
            pass
        elif dept_students:
            query["registerNumber"] = {"$in": dept_students}
        elif department and department.upper() != "ALL":
            return {"count": 0}

    count = await db.student_subjects.count_documents(query)
    return {"count": count}

@router.delete("/previous-results")
async def delete_previous_results(
    department: Optional[str] = Query(None),
    resultSem: Optional[int] = Query(None),
    regulation: Optional[str] = Query(None),
    registerNumber: Optional[str] = Query(None)
):
    query = {}
    if resultSem is not None and resultSem != 0:
        query["semester"] = int(resultSem)
    if registerNumber:
        query["registerNumber"] = registerNumber.strip().upper()

    student_query = {}
    if department and department.upper() != "ALL":
        student_query["department"] = department.strip().upper()

    if student_query or (department and department.upper() != "ALL"):
        dept_students = [
            doc["registerNumber"]
            async for doc in db.students.find(student_query, {"registerNumber": 1})
        ]
        if registerNumber and registerNumber.strip().upper() not in dept_students:
            dept_students.append(registerNumber.strip().upper())

        if query.get("registerNumber"):
            pass
        elif dept_students:
            query["registerNumber"] = {"$in": dept_students}
        elif department and department.upper() != "ALL":
            return {"message": f"No student records found for department {department}.", "deletedCount": 0}

    if not query:
        raise HTTPException(
            status_code=400,
            detail="Please specify at least department, resultSem, or registerNumber to delete."
        )

    res = await db.student_subjects.delete_many(query)
    return {
        "message": f"Successfully deleted {res.deleted_count} previous result record(s).",
        "deletedCount": res.deleted_count
    }



