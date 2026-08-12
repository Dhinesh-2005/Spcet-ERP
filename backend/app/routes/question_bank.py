import re
import random
import base64
from datetime import datetime
from io import BytesIO
from typing import Optional, List, Dict, Any
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Query
from bson import ObjectId
from app.database import db, clean_doc, clean_docs
from app.models.schemas import QuestionBankModel, GenerateQuestionPaperRequest

router = APIRouter()

def _clean_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()

def _find_col_idx(header: List[Any], keywords: List[str]) -> int:
    for idx, cell in enumerate(header):
        val = re.sub(r"[^a-z0-9]", "", _clean_str(cell).lower())
        for kw in keywords:
            if kw in val:
                return idx
    return -1

import zipfile
import xml.etree.ElementTree as ET

def _extract_all_excel_images(content_bytes, sheet=None):
    image_map = {}

    def _detect_fmt(raw):
        if raw[0:2] == b"\xff\xd8":
            return "jpeg"
        if raw[0:4] == b"\x89PNG":
            return "png"
        if raw[0:3] == b"GIF":
            return "gif"
        if raw[0:2] == b"BM":
            return "bmp"
        if raw[0:4] == b"RIFF" and raw[8:12] == b"WEBP":
            return "webp"
        return "png"

    def _img_obj(raw):
        if not raw:
            return None
        fmt = _detect_fmt(raw)
        b64 = base64.b64encode(raw).decode("utf-8")
        return {"base64": "data:image/" + fmt + ";base64," + b64}

    def _resolve_path(base_dir, target, file_list):
        if target.startswith("../"):
            parent = base_dir.rsplit("/", 1)[0] if "/" in base_dir else ""
            resolved = (parent + "/" + target[3:]).lstrip("/")
        elif target.startswith("/"):
            resolved = target.lstrip("/")
        else:
            resolved = (base_dir + "/" + target).lstrip("/")
        resolved = "/".join(p for p in resolved.split("/") if p)
        if resolved not in file_list:
            rl = resolved.lower()
            resolved = next((f for f in file_list if f.lower() == rl), resolved)
        return resolved

    try:
        zf = zipfile.ZipFile(BytesIO(content_bytes))
        file_list = zf.namelist()

        # ---------------------------------------------------------------
        # METHOD 1: Modern Excel "in-cell image" via xl/richData
        # richValueRel.xml lists rId -> image relationship (index = order)
        # Sheet cells have vm="N" attribute (1-based) -> richvalue index
        # rdrichvalue.xml maps richvalue index -> rvRel index -> image
        # ---------------------------------------------------------------
        rvr_file = "xl/richData/richValueRel.xml"
        rvr_rels_file = "xl/richData/_rels/richValueRel.xml.rels"
        if rvr_file in file_list and rvr_rels_file in file_list:
            try:
                # Build rId -> image path from rels
                rid_to_img = {}
                for child in ET.fromstring(zf.read(rvr_rels_file)):
                    rid = child.attrib.get("Id", "")
                    tgt = child.attrib.get("Target", "")
                    resolved = _resolve_path("xl/richData", tgt, file_list)
                    if rid and resolved in file_list:
                        rid_to_img[rid] = resolved

                # Build ordered list of image paths from richValueRel.xml
                rvr_ordered = []
                rvr_tree = ET.fromstring(zf.read(rvr_file))
                for rel in rvr_tree:
                    rid = None
                    for attr_k, attr_v in rel.attrib.items():
                        if attr_k.endswith("}id") or attr_k == "id":
                            rid = attr_v
                            break
                    if rid and rid in rid_to_img:
                        rvr_ordered.append(rid_to_img[rid])
                    else:
                        rvr_ordered.append(None)

                # Parse rdrichvalue.xml: each <rv> element's first <v> is the rvRel index (0-based)
                # vm index in cell is 1-based index into rdrichvalue entries
                rv_file = "xl/richData/rdrichvalue.xml"
                vm_to_img = {}  # vm value (int) -> image path
                if rv_file in file_list:
                    rv_tree = ET.fromstring(zf.read(rv_file))
                    rv_entries = [e for e in rv_tree if e.tag.split("}")[-1] == "rv"]
                    for rv_idx, rv_elem in enumerate(rv_entries):
                        vals = [c.text for c in rv_elem if c.tag.split("}")[-1] == "v"]
                        if vals:
                            try:
                                rvrel_idx = int(vals[0])
                                if 0 <= rvrel_idx < len(rvr_ordered) and rvr_ordered[rvrel_idx]:
                                    vm_to_img[rv_idx + 1] = rvr_ordered[rvrel_idx]
                            except (ValueError, TypeError):
                                pass
                else:
                    # Fallback: vm=1 -> first image
                    if rvr_ordered and rvr_ordered[0]:
                        vm_to_img[1] = rvr_ordered[0]

                # Pre-load images to avoid re-reading bytes multiple times
                img_cache = {}

                # Parse sheet XML to find cells with vm attribute
                sheet_files = [f for f in file_list if re.match(r"xl/worksheets/sheet\d+\.xml$", f)]
                for sf in sheet_files:
                    try:
                        s_tree = ET.fromstring(zf.read(sf))
                        ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
                        for row_elem in s_tree.iter("{" + ns + "}row"):
                            # Row number is 1-based in XML; convert to 0-based for image_map
                            row_r = row_elem.attrib.get("r")
                            row_0 = int(row_r) - 1 if row_r else None
                            for c_elem in row_elem:
                                vm_str = c_elem.attrib.get("vm")
                                if vm_str is None:
                                    continue
                                try:
                                    vm_val = int(vm_str)
                                except ValueError:
                                    continue
                                cell_ref = c_elem.attrib.get("r", "")
                                # Parse column from cell ref e.g. "C2" -> col index 2 (0-based)
                                col_letters = "".join(ch for ch in cell_ref if ch.isalpha())
                                col_0 = 0
                                for ch in col_letters:
                                    col_0 = col_0 * 26 + (ord(ch.upper()) - ord("A") + 1)
                                col_0 -= 1  # 0-based

                                img_path = vm_to_img.get(vm_val)
                                if not img_path:
                                    continue
                                if img_path not in img_cache:
                                    raw = zf.read(img_path)
                                    img_cache[img_path] = _img_obj(raw)
                                img_obj = img_cache[img_path]
                                if img_obj and row_0 is not None:
                                    if row_0 not in image_map:
                                        image_map[row_0] = img_obj
                                    image_map[(row_0, col_0)] = img_obj
                    except Exception:
                        pass
            except Exception:
                pass

        # ---------------------------------------------------------------
        # METHOD 2: Classic floating drawing images (twoCellAnchor)
        # ---------------------------------------------------------------
        drawing_files = [f for f in file_list if re.search(r"drawings/drawing\d*\.xml$", f)]
        for dw_file in drawing_files:
            parts = dw_file.rsplit("/", 1)
            dw_dir = parts[0] if len(parts) == 2 else ""
            dw_name = parts[-1]
            rel_file = dw_dir + "/_rels/" + dw_name + ".rels" if dw_dir else "_rels/" + dw_name + ".rels"

            rel_map = {}
            if rel_file in file_list:
                try:
                    for child in ET.fromstring(zf.read(rel_file)):
                        rid = child.attrib.get("Id", "")
                        tgt = child.attrib.get("Target", "")
                        if rid and tgt:
                            rel_map[rid] = _resolve_path(dw_dir, tgt, file_list)
                except Exception:
                    pass

            if not rel_map:
                continue

            try:
                dw_tree = ET.fromstring(zf.read(dw_file))
                for anchor in dw_tree.iter():
                    tag = anchor.tag.split("}")[-1] if "}" in anchor.tag else anchor.tag
                    if tag not in ("twoCellAnchor", "oneCellAnchor"):
                        continue
                    from_elem = next((c for c in anchor if c.tag.split("}")[-1] == "from"), None)
                    if from_elem is None:
                        continue
                    row_idx = col_idx = None
                    for c in from_elem:
                        t = c.tag.split("}")[-1] if "}" in c.tag else c.tag
                        if t == "row" and c.text:
                            try: row_idx = int(c.text)
                            except ValueError: pass
                        elif t == "col" and c.text:
                            try: col_idx = int(c.text)
                            except ValueError: pass
                    if row_idx is None:
                        continue
                    r_id = None
                    for blip in anchor.iter():
                        if blip.tag.split("}")[-1] == "blip":
                            for k, v in blip.attrib.items():
                                if k.split("}")[-1] == "embed":
                                    r_id = v
                                    break
                            if r_id:
                                break
                    if not r_id or r_id not in rel_map:
                        continue
                    img_path = rel_map[r_id]
                    if img_path in file_list:
                        raw = zf.read(img_path)
                        img_obj = _img_obj(raw)
                        if img_obj:
                            if row_idx not in image_map:
                                image_map[row_idx] = img_obj
                            if col_idx is not None:
                                image_map[(row_idx, col_idx)] = img_obj
            except Exception:
                pass
    except Exception:
        pass

    # ---------------------------------------------------------------
    # METHOD 3: openpyxl sheet._images fallback
    # ---------------------------------------------------------------
    if sheet and hasattr(sheet, "_images") and sheet._images:
        for img in sheet._images:
            try:
                row_idx = col_idx = None
                anchor = getattr(img, "anchor", None)
                if anchor:
                    if hasattr(anchor, "_from"):
                        row_idx = anchor._from.row
                        col_idx = anchor._from.col
                    elif hasattr(anchor, "row"):
                        row_idx = anchor.row
                        col_idx = getattr(anchor, "col", None)
                if row_idx is None:
                    continue
                raw = None
                if hasattr(img, "_data"):
                    raw = img._data()
                elif hasattr(img, "ref"):
                    ref = img.ref
                    raw = ref.read() if hasattr(ref, "read") else (ref if isinstance(ref, bytes) else None)
                elif hasattr(img, "path"):
                    with open(img.path, "rb") as f:
                        raw = f.read()
                if not raw and hasattr(img, "image"):
                    try:
                        buf = BytesIO()
                        img.image.save(buf, format="PNG")
                        raw = buf.getvalue()
                    except Exception:
                        pass
                if raw:
                    img_obj = _img_obj(raw)
                    if img_obj:
                        if row_idx not in image_map:
                            image_map[row_idx] = img_obj
                        if col_idx is not None and (row_idx, col_idx) not in image_map:
                            image_map[(row_idx, col_idx)] = img_obj
            except Exception:
                pass

    return image_map


@router.post("/debug-images")
async def debug_image_extraction(file: UploadFile = File(...)):
    """Debug endpoint: returns image_map keys extracted from the uploaded Excel."""
    import openpyxl
    content = await file.read()
    try:
        workbook = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    # Detect header row
    header_idx = 0
    for idx, row in enumerate(rows[:20]):
        if not row:
            continue
        row_str = " ".join([_clean_str(c).lower() for c in row if c])
        if "question" in row_str or "co" in row_str or "part" in row_str:
            header_idx = idx
            break

    header = [_clean_str(c) for c in (rows[header_idx] or [])]
    img_idx = _find_col_idx(header, ["image", "img", "diagram", "figure", "picture"])
    image_map = _extract_all_excel_images(content, sheet)

    # Serialize keys (tuples are not JSON serialisable)
    keys_info = []
    for k in image_map.keys():
        if isinstance(k, tuple):
            keys_info.append({"type": "tuple", "row": k[0], "col": k[1]})
        else:
            keys_info.append({"type": "row_only", "row": k})

    # Also dump what zip contains
    try:
        import zipfile as zf_mod
        zf = zf_mod.ZipFile(BytesIO(content))
        zip_files = zf.namelist()
    except Exception:
        zip_files = []

    return {
        "header_idx": header_idx,
        "header": header,
        "img_col_idx": img_idx,
        "total_data_rows": len(rows) - header_idx - 1,
        "image_map_keys": keys_info,
        "image_count": len([k for k in image_map if isinstance(k, int)]),
        "zip_drawing_files": [f for f in zip_files if "drawing" in f.lower()],
        "zip_media_files": [f for f in zip_files if "media" in f.lower()],
    }

def _normalize_unit(unit_str: str, co_str: str = "") -> str:
    unit_str = _clean_str(unit_str)
    if unit_str:
        num_match = re.search(r"\d+", unit_str)
        if num_match:
            return f"Unit {num_match.group(0)}"
        return unit_str
    
    co_str = _clean_str(co_str).upper()
    co_match = re.search(r"CO(\d+)", co_str)
    if co_match:
        return f"Unit {co_match.group(1)}"
    
    return "Unit 1"

def _normalize_co(co_str: str) -> str:
    co_str = _clean_str(co_str).upper()
    match = re.search(r"CO\d+", co_str)
    if match:
        return match.group(0)
    num_match = re.search(r"\d+", co_str)
    if num_match:
        return f"CO{num_match.group(0)}"
    return co_str if co_str else "CO1"

def process_question_equation(question_text: str):
    """
    Detects if question contains an equation/math formula,
    converts mathematical patterns to LaTeX/KaTeX, protects CO1-CO6 labels,
    and returns a structured dict with normal text and equation parts.
    """
    if not question_text:
        return {"question": "", "equation": "", "hasEquation": False, "latex": ""}

    raw = str(question_text).strip()
    
    # 1. Protect CO tokens (CO1 - CO99) so they are NEVER converted to subscripts or equations
    co_placeholders = {}
    def _preserve_co(match):
        token = match.group(0)
        placeholder = f"__CO_TOKEN_{len(co_placeholders)}__"
        co_placeholders[placeholder] = token
        return placeholder

    protected_text = re.sub(r"\bCO[0-9]+\b", _preserve_co, raw)

    # 2. Check for mathematical / chemical equation patterns
    equation_symbols = [
        r"=", r"√", r"∫", r"Σ", r"π", r"α", r"β", r"θ", r"λ", r"\^",
        r"²", r"³", r"⁻", r"±", r"≠", r"≤", r"≥", r"dy/dx", r"d/dx"
    ]
    
    has_eq_symbol = any(re.search(pat, protected_text) for pat in equation_symbols)
    has_exponent_pattern = bool(re.search(r"\b[a-zA-Z](\d+)\b", protected_text))
    chem_matches = [m.group(0) for m in re.finditer(r"\b[A-Z][a-z]?\d+(?:[A-Z][a-z]?\d*)*\b", protected_text) if "__CO_TOKEN_" not in m.group(0)]
    has_chem_pattern = len(chem_matches) > 0

    is_equation = has_eq_symbol or has_exponent_pattern or has_chem_pattern

    if not is_equation:
        final_text = protected_text
        for ph, orig in co_placeholders.items():
            final_text = final_text.replace(ph, orig)
        return {
            "question": final_text,
            "equation": "",
            "hasEquation": False,
            "latex": ""
        }

    # 3. Separate text prefix from equation part
    text_part = ""
    eq_part = ""

    triggers = [
        r"calculate the value of", r"calculate the area using", r"find the value of",
        r"find the area using", r"solve the equation", r"prove that", r"the value of",
        r"calculate", r"determine", r"evaluate", r"solve", r"find", r"using", r"where", r"what is"
    ]

    found_split = False
    for trig in triggers:
        match = re.search(r"^(.*?\b" + trig + r"\s+)(.+)$", protected_text, re.IGNORECASE)
        if match:
            rest = match.group(2).strip()
            if any(re.search(pat, rest) for pat in equation_symbols) or re.search(r"\b[a-zA-Z]\d+\b", rest) or len(chem_matches) > 0:
                text_part = match.group(1).strip()
                eq_part = rest
                found_split = True
                break

    if not found_split:
        if "=" in protected_text:
            idx = protected_text.find("=")
            before_eq = protected_text[:idx]
            match_start = re.search(r"^(.*?)([a-zA-Z0-9_\(\)\s\+\-\*\^²/√∫Σπ]+)$", before_eq)
            if match_start and match_start.group(1).strip():
                text_part = match_start.group(1).strip()
                eq_part = (match_start.group(2) + protected_text[idx:]).strip()
            else:
                eq_part = protected_text.strip()
        else:
            eq_part = protected_text.strip()

    # 4. Convert eq_part to LaTeX
    converted_eq = eq_part

    # Exponents x2 -> x^2, y3 -> y^3
    converted_eq = re.sub(r"\b([a-zA-Z])(\d+)\b", r"\1^\2", converted_eq)
    converted_eq = re.sub(r"²", "^2", converted_eq)
    converted_eq = re.sub(r"³", "^3", converted_eq)
    converted_eq = re.sub(r"10⁻²", "10^{-2}", converted_eq)
    converted_eq = re.sub(r"10\^-2", "10^{-2}", converted_eq)

    # Math symbols
    converted_eq = converted_eq.replace("√", r"\sqrt")
    converted_eq = converted_eq.replace("∫", r"\int ")
    converted_eq = converted_eq.replace("Σ", r"\sum ")
    converted_eq = converted_eq.replace("π", r"\pi ")
    converted_eq = converted_eq.replace("α", r"\alpha ")
    converted_eq = converted_eq.replace("β", r"\beta ")
    converted_eq = converted_eq.replace("θ", r"\theta ")
    converted_eq = converted_eq.replace("λ", r"\lambda ")

    # Chemical formula conversion for non-CO tokens
    def _convert_chem(match):
        formula = match.group(0)
        if "__CO_TOKEN_" in formula:
            return formula
        return re.sub(r"([A-Za-z])(\d+)", r"\1_\2", formula)

    converted_eq = re.sub(r"\b[A-Z][a-z]?\d+(?:[A-Z][a-z]?\d*)*\b", _convert_chem, converted_eq)

    # Restore protected CO tokens
    for ph, orig in co_placeholders.items():
        text_part = text_part.replace(ph, orig)
        converted_eq = converted_eq.replace(ph, orig)
    display_question = raw

    return {
        "question": display_question,
        "equation": converted_eq,
        "hasEquation": True if converted_eq else False,
        "latex": f"{text_part} ${converted_eq}$".strip() if text_part else f"${converted_eq}$"
    }

def _normalize_part(part_str: str, marks_val: float) -> str:
    part_str = _clean_str(part_str).upper()
    if "A" in part_str:
        return "A"
    if "B" in part_str:
        return "B"
    if "C" in part_str:
        return "C"
    
    if marks_val <= 2:
        return "A"
    elif marks_val >= 5:
        return "B"
    return "A"

@router.post("/upload")
async def upload_question_bank(
    file: UploadFile = File(...),
    subjectCode: str = Form(...),
    subjectName: Optional[str] = Form(None),
    regulation: Optional[str] = Form(None),
    semester: Optional[str] = Form(None),
):
    import openpyxl

    content = await file.read()
    try:
        workbook = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    header_idx = -1
    for idx, row in enumerate(rows[:20]):
        if not row:
            continue
        row_str = " ".join([_clean_str(c).lower() for c in row if c])
        if "question" in row_str or "co" in row_str or "part" in row_str:
            header_idx = idx
            break

    if header_idx == -1:
        header_idx = 0

    header = [ _clean_str(c) for c in (rows[header_idx] or []) ]
    q_idx = _find_col_idx(header, ["question", "qtext", "questions"])
    co_idx = _find_col_idx(header, ["co", "courseoutcome", "c/o", "co/po"])
    k_idx = _find_col_idx(header, ["klevel", "k-level", "btl", "bloom"])
    marks_idx = _find_col_idx(header, ["mark", "marks", "weightage"])
    part_idx = _find_col_idx(header, ["part", "section"])
    unit_idx = _find_col_idx(header, ["unit", "unitno"])
    img_idx = _find_col_idx(header, ["image", "img", "diagram", "figure", "picture"])

    if q_idx == -1:
        raise HTTPException(status_code=400, detail="Missing mandatory 'Question' column in Excel file")

    image_map = _extract_all_excel_images(content, sheet)

    total_rows = 0
    imported = 0
    skipped = 0
    duplicates = 0
    failed = 0

    now_iso = datetime.utcnow().isoformat()

    existing_docs = [doc async for doc in db.QuestionBank.find({"subjectCode": subjectCode})]
    existing_map = set(re.sub(r"\s+", "", doc.get("question", "").lower()) for doc in existing_docs if doc.get("question"))

    new_docs = []

    for row_num in range(header_idx + 1, len(rows)):
        row = rows[row_num]
        if not row:
            continue

        total_rows += 1

        question_text = _clean_str(row[q_idx]) if q_idx < len(row) else ""
        if not question_text:
            skipped += 1
            continue

        raw_co = _clean_str(row[co_idx]) if co_idx != -1 and co_idx < len(row) else "CO1"
        co_val = _normalize_co(raw_co)

        k_level = _clean_str(row[k_idx]) if k_idx != -1 and k_idx < len(row) else "K1"
        if not k_level:
            k_level = "K1"

        raw_marks = row[marks_idx] if marks_idx != -1 and marks_idx < len(row) else None
        try:
            marks_val = float(raw_marks) if raw_marks is not None and str(raw_marks).strip() != "" else 2.0
        except (ValueError, TypeError):
            marks_val = 2.0

        raw_part = _clean_str(row[part_idx]) if part_idx != -1 and part_idx < len(row) else ""
        part_val = _normalize_part(raw_part, marks_val)

        raw_unit = _clean_str(row[unit_idx]) if unit_idx != -1 and unit_idx < len(row) else ""
        unit_val = _normalize_unit(raw_unit, co_val)

        img_data = image_map.get((row_num, img_idx)) if img_idx != -1 else None
        if not img_data:
            img_data = image_map.get(row_num)
        if not img_data and img_idx != -1 and img_idx < len(row):
            cell_img_val = _clean_str(row[img_idx])
            if cell_img_val.startswith("data:image") or cell_img_val.startswith("http"):
                img_data = {"base64": cell_img_val}

        norm_q = re.sub(r"\s+", "", question_text.lower())
        if norm_q in existing_map:
            duplicates += 1
            continue

        existing_map.add(norm_q)

        processed_eq = process_question_equation(question_text)
        doc = {
            "subjectCode": subjectCode,
            "subjectName": subjectName,
            "regulation": regulation,
            "semester": semester,
            "unit": unit_val,
            "question": processed_eq["question"],
            "equation": processed_eq["equation"],
            "hasEquation": processed_eq["hasEquation"],
            "latex": processed_eq["latex"],
            "part": part_val,
            "marks": marks_val,
            "co": co_val,
            "kLevel": k_level,
            "image": img_data,
            "createdAt": now_iso,
            "updatedAt": now_iso,
        }
        new_docs.append(doc)
        imported += 1

    if new_docs:
        await db.QuestionBank.insert_many(new_docs)

    return {
        "message": f"Question Bank upload completed for {subjectCode}",
        "summary": {
            "totalRows": total_rows,
            "imported": imported,
            "skipped": skipped,
            "duplicates": duplicates,
            "failed": failed,
        }
    }

@router.post("/parse-and-generate")
async def parse_and_generate_paper(
    file: UploadFile = File(...),
    unit: str = Form("Unit 1"),
):
    import openpyxl

    content = await file.read()
    try:
        workbook = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    header_idx = -1
    for idx, row in enumerate(rows[:20]):
        if not row:
            continue
        row_str = " ".join([_clean_str(c).lower() for c in row if c])
        if "question" in row_str or "co" in row_str or "part" in row_str:
            header_idx = idx
            break

    if header_idx == -1:
        header_idx = 0

    header = [ _clean_str(c) for c in (rows[header_idx] or []) ]
    q_idx = _find_col_idx(header, ["question", "qtext", "questions"])
    co_idx = _find_col_idx(header, ["co", "courseoutcome", "c/o", "co/po"])
    k_idx = _find_col_idx(header, ["klevel", "k-level", "btl", "bloom"])
    marks_idx = _find_col_idx(header, ["mark", "marks", "weightage"])
    part_idx = _find_col_idx(header, ["part", "section"])
    unit_idx = _find_col_idx(header, ["unit", "unitno"])
    img_idx = _find_col_idx(header, ["image", "img", "diagram", "figure", "picture"])

    if q_idx == -1:
        raise HTTPException(status_code=400, detail="Missing mandatory 'Question' column in Excel file")

    image_map = _extract_all_excel_images(content, sheet)

    unit_num_match = re.search(r"\d+", unit)
    target_unit_num = unit_num_match.group(0) if unit_num_match else None

    all_questions = []
    seen_q = set()

    for row_num in range(header_idx + 1, len(rows)):
        row = rows[row_num]
        if not row:
            continue

        question_text = _clean_str(row[q_idx]) if q_idx < len(row) else ""
        if not question_text:
            continue

        norm_q = re.sub(r"\s+", "", question_text.lower())
        if norm_q in seen_q:
            continue
        seen_q.add(norm_q)

        raw_co = _clean_str(row[co_idx]) if co_idx != -1 and co_idx < len(row) else ""
        if raw_co:
            co_val = _normalize_co(raw_co)
        else:
            co_val = f"CO{target_unit_num}" if target_unit_num else "CO1"

        k_level = _clean_str(row[k_idx]) if k_idx != -1 and k_idx < len(row) else "K1"
        if not k_level:
            k_level = "K1"

        raw_marks = row[marks_idx] if marks_idx != -1 and marks_idx < len(row) else None
        try:
            marks_val = float(raw_marks) if raw_marks is not None and str(raw_marks).strip() != "" else 2.0
        except (ValueError, TypeError):
            marks_val = 2.0

        raw_part = _clean_str(row[part_idx]) if part_idx != -1 and part_idx < len(row) else ""
        part_val = _normalize_part(raw_part, marks_val)

        raw_unit = _clean_str(row[unit_idx]) if unit_idx != -1 and unit_idx < len(row) else ""
        unit_val = _normalize_unit(raw_unit, co_val)

        img_data = image_map.get((row_num, img_idx)) if img_idx != -1 else None
        if not img_data:
            img_data = image_map.get(row_num)
        if not img_data and img_idx != -1 and img_idx < len(row):
            cell_img_val = _clean_str(row[img_idx])
            if cell_img_val.startswith("data:image") or cell_img_val.startswith("http"):
                img_data = {"base64": cell_img_val}

        # Filter by unit if specified
        if target_unit_num:
            u_match = re.search(r"\d+", unit_val)
            if u_match and u_match.group(0) != target_unit_num:
                continue

        processed_eq = process_question_equation(question_text)
        doc = {
            "unit": unit_val,
            "question": processed_eq["question"],
            "equation": processed_eq["equation"],
            "hasEquation": processed_eq["hasEquation"],
            "latex": processed_eq["latex"],
            "part": part_val,
            "marks": marks_val,
            "co": co_val,
            "kLevel": k_level,
            "image": img_data,
        }
        all_questions.append(doc)

    part_a = [q for q in all_questions if q["part"] == "A" or q["marks"] <= 2]
    part_b_all = [q for q in all_questions if q["part"] == "B" or q["marks"] > 2]
    part_c_all = [q for q in all_questions if q["part"] == "C"]

    b_16 = [q for q in part_b_all if q["marks"] >= 12]
    b_8 = [q for q in part_b_all if 5 <= q["marks"] < 12]

    random.shuffle(part_a)
    random.shuffle(part_b_all)
    random.shuffle(b_16)
    random.shuffle(b_8)
    random.shuffle(part_c_all)

    # For Regulation 2024:
    # Q6 (16m): 2 questions (Q6a, Q6b)
    # Q7 (16m): 2 questions (Q7a, Q7b)
    # Q8 (8m):  2 questions (Q8a, Q8b)
    selected_16 = b_16[:4]
    if len(selected_16) < 4:
        used = {id(q) for q in selected_16}
        rem = [q for q in part_b_all if id(q) not in used]
        selected_16.extend(rem[: 4 - len(selected_16)])

    used_16 = {id(q) for q in selected_16}
    selected_8 = [q for q in b_8 if id(q) not in used_16][:2]
    if len(selected_8) < 2:
        rem = [q for q in part_b_all if id(q) not in used_16]
        selected_8.extend(rem[: 2 - len(selected_8)])

    for q in selected_16:
        q["marks"] = 16
    for q in selected_8:
        q["marks"] = 8

    selected_part_b = selected_16 + selected_8
    selected_part_a = part_a[:5]

    if len(part_c_all) >= 2:
        selected_part_c = part_c_all[:2]
    else:
        used_b = {id(q) for q in selected_part_b}
        leftover_b = [q for q in part_b_all if id(q) not in used_b]
        selected_part_c = (part_c_all + leftover_b)[:2]

    warning = None
    if len(part_a) < 5 or len(part_b_all) < 6:
        warning = f"Notice: Excel provided {len(part_a)} Part A and {len(part_b_all)} Part B questions for {unit}."

    return {
        "unit": unit,
        "warning": warning,
        "partA": selected_part_a,
        "partB": selected_part_b,
        "partC": selected_part_c,
    }

@router.get("")
async def get_questions(
    subjectCode: Optional[str] = Query(None),
    unit: Optional[str] = Query(None),
    co: Optional[str] = Query(None),
    part: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
):
    query: Dict[str, Any] = {}
    if subjectCode:
        query["subjectCode"] = subjectCode
    if unit:
        num = re.search(r"\d+", unit)
        if num:
            query["unit"] = {"$regex": f"Unit\\s*{num.group(0)}|\\b{num.group(0)}\\b", "$options": "i"}
        else:
            query["unit"] = unit
    if co:
        query["co"] = co
    if part:
        query["part"] = part.upper()
    if search:
        query["question"] = {"$regex": search, "$options": "i"}

    total = await db.QuestionBank.count_documents(query)
    skip = (page - 1) * limit
    cursor = db.QuestionBank.find(query).skip(skip).limit(limit).sort("createdAt", -1)
    items = [clean_doc(d) async for d in cursor]

    pages = (total + limit - 1) // limit if total > 0 else 1

    return {
        "questions": items,
        "total": total,
        "page": page,
        "pages": pages,
        "limit": limit,
    }

@router.post("/manual")
async def add_question(question: QuestionBankModel):
    now_iso = datetime.utcnow().isoformat()
    doc = question.dict(exclude_none=True, exclude={"id"})
    doc["createdAt"] = now_iso
    doc["updatedAt"] = now_iso
    doc["unit"] = _normalize_unit(doc.get("unit", ""), doc.get("co", ""))
    doc["co"] = _normalize_co(doc.get("co", ""))
    doc["part"] = _normalize_part(doc.get("part", ""), doc.get("marks", 2.0))

    res = await db.QuestionBank.insert_one(doc)
    doc["id"] = str(res.inserted_id)
    return {"message": "Question created successfully", "question": doc}

@router.put("/{question_id}")
async def update_question(question_id: str, question: QuestionBankModel):
    now_iso = datetime.utcnow().isoformat()
    update_data = question.dict(exclude_none=True, exclude={"id"})
    update_data["updatedAt"] = now_iso
    if "unit" in update_data:
        update_data["unit"] = _normalize_unit(update_data["unit"], update_data.get("co", ""))
    if "co" in update_data:
        update_data["co"] = _normalize_co(update_data["co"])
    if "part" in update_data:
        update_data["part"] = _normalize_part(update_data["part"], update_data.get("marks", 2.0))

    result = await db.QuestionBank.update_one({"_id": ObjectId(question_id)}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Question not found")
    return {"message": "Question updated successfully"}

@router.delete("/{question_id}")
async def delete_question(question_id: str):
    result = await db.QuestionBank.delete_one({"_id": ObjectId(question_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Question not found")
    return {"message": "Question deleted successfully"}

@router.post("/generate-paper")
async def generate_question_paper(req: GenerateQuestionPaperRequest):
    unit_num_match = re.search(r"\d+", req.unit)
    unit_regex = f"Unit\\s*{unit_num_match.group(0)}|\\b{unit_num_match.group(0)}\\b" if unit_num_match else req.unit

    base_query = {
        "subjectCode": req.subjectCode,
        "unit": {"$regex": unit_regex, "$options": "i"},
    }

    part_a_query = {**base_query, "part": "A"}
    part_b_query = {**base_query, "part": "B"}
    part_c_query = {**base_query, "part": "C"}

    part_a_all = [clean_doc(d) async for d in db.QuestionBank.find(part_a_query)]
    part_b_all = [clean_doc(d) async for d in db.QuestionBank.find(part_b_query)]
    part_c_all = [clean_doc(d) async for d in db.QuestionBank.find(part_c_query)]

    random.shuffle(part_a_all)
    random.shuffle(part_b_all)
    random.shuffle(part_c_all)

    selected_part_a = part_a_all[:req.partACount]
    
    if req.partBCount == 6:
        b_16 = [q for q in part_b_all if q.get("marks", 16) >= 12]
        b_8 = [q for q in part_b_all if 5 <= q.get("marks", 8) < 12]
        random.shuffle(b_16)
        random.shuffle(b_8)

        selected_16 = b_16[:4]
        if len(selected_16) < 4:
            used = {id(q) for q in selected_16}
            rem = [q for q in part_b_all if id(q) not in used]
            selected_16.extend(rem[: 4 - len(selected_16)])

        used_16 = {id(q) for q in selected_16}
        selected_8 = [q for q in b_8 if id(q) not in used_16][:2]
        if len(selected_8) < 2:
            rem = [q for q in part_b_all if id(q) not in used_16]
            selected_8.extend(rem[: 2 - len(selected_8)])

        for q in selected_16:
            q["marks"] = 16
        for q in selected_8:
            q["marks"] = 8

        selected_part_b = selected_16 + selected_8
    else:
        selected_part_b = part_b_all[:req.partBCount]

    # If Part C doesn't have dedicated Part C questions, fallback to leftover Part B questions
    if len(part_c_all) >= req.partCCount:
        selected_part_c = part_c_all[:req.partCCount]
    else:
        leftover_b = part_b_all[req.partBCount:]
        selected_part_c = (part_c_all + leftover_b)[:req.partCCount]

    warning = None
    if len(part_a_all) < req.partACount or len(part_b_all) < req.partBCount or (len(part_c_all) < req.partCCount and len(part_b_all) < req.partBCount + req.partCCount):
        warning = f"Warning: Found {len(part_a_all)} Part A, {len(part_b_all)} Part B, and {len(part_c_all)} Part C questions for {req.subjectCode} {req.unit}."

    now_iso = datetime.utcnow().isoformat()
    history_entry = {
        "subjectCode": req.subjectCode,
        "unit": req.unit,
        "generatedAt": now_iso,
        "partACount": len(selected_part_a),
        "partBCount": len(selected_part_b),
        "partCCount": len(selected_part_c),
        "warning": warning,
    }
    await db.question_paper_history.insert_one(history_entry)

    return {
        "subjectCode": req.subjectCode,
        "unit": req.unit,
        "warning": warning,
        "partA": selected_part_a,
        "partB": selected_part_b,
        "partC": selected_part_c,
    }


@router.get("/paper-history")
async def get_paper_history():
    cursor = db.question_paper_history.find({}).sort("generatedAt", -1).limit(50)
    return [clean_doc(d) async for d in cursor]
